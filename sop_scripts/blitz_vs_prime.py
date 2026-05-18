from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ======== COLORS ========
PURPLE = "5B2D8E"
DARK_PURPLE = "3D1D5E"
WHITE = "FFFFFF"
LIGHT_PURPLE = "F5F0FA"
BLITZ_ORANGE = "E67E22"
BLITZ_LIGHT = "FDF2E9"
PRIME_BLUE = "2980B9"
PRIME_LIGHT = "EBF5FB"
GREEN = "27AE60"
GREEN_LIGHT = "EAFAF1"
GRAY = "BDC3C7"
GRAY_LIGHT = "F8F9FA"
DARK = "2C3E50"
RED = "E74C3C"
YELLOW_BG = "FFF9C4"
SHARED_GREEN = "1ABC9C"
SHARED_LIGHT = "E8F8F5"

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill_purple = PatternFill("solid", fgColor=PURPLE)
hdr_fill_blitz = PatternFill("solid", fgColor=BLITZ_ORANGE)
hdr_fill_prime = PatternFill("solid", fgColor=PRIME_BLUE)
hdr_fill_shared = PatternFill("solid", fgColor=SHARED_GREEN)
hdr_fill_dark = PatternFill("solid", fgColor=DARK)
title_font = Font(name="Arial", bold=True, color=PURPLE, size=16)
subtitle_font = Font(name="Arial", bold=True, color=DARK, size=12)
bold_font = Font(name="Arial", bold=True, size=10)
normal_font = Font(name="Arial", size=10)
small_font = Font(name="Arial", size=9, color="666666")
blitz_font = Font(name="Arial", bold=True, color=BLITZ_ORANGE, size=10)
prime_font = Font(name="Arial", bold=True, color=PRIME_BLUE, size=10)
money_font = Font(name="Arial", bold=True, color=GREEN, size=11)
wrap = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))

def style_header_row(ws, row, cols, fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = fill
        cell.alignment = center_wrap
        cell.border = thin_border

def style_data_cell(ws, row, col, font=None, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font or normal_font
    cell.alignment = wrap
    cell.border = thin_border
    if fill: cell.fill = PatternFill("solid", fgColor=fill)
    return cell

# ======================================================================
# SHEET 1: SIDE-BY-SIDE TIMELINE
# ======================================================================
ws1 = wb.active
ws1.title = "Timeline Comparison"
ws1.sheet_properties.tabColor = PURPLE

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 34
ws1.column_dimensions['C'].width = 34
ws1.column_dimensions['D'].width = 34
ws1.column_dimensions['E'].width = 16

# Title
ws1.merge_cells('A1:E1')
ws1['A1'].font = title_font
ws1['A1'].value = "APW — THE BLITZ vs THE PRIME PROTOCOL"
ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:E2')
ws1['A2'].font = Font(name="Arial", italic=True, color="666666", size=10)
ws1['A2'].value = "Side-by-Side Timeline — Week by Week Execution Roadmap"
ws1['A2'].alignment = Alignment(horizontal="center")

# Headers Row 4
r = 4
headers = ["TIMING", "THE BLITZ (V1)", "THE PRIME PROTOCOL (V2)", "SHARED ACTIONS", "MODULE"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
fills = [hdr_fill_purple, hdr_fill_blitz, hdr_fill_prime, hdr_fill_shared, hdr_fill_dark]
for i, f in enumerate(fills, 1):
    ws1.cell(row=r, column=i).fill = f
    ws1.cell(row=r, column=i).font = hdr_font
    ws1.cell(row=r, column=i).alignment = center_wrap
    ws1.cell(row=r, column=i).border = thin_border
ws1.row_dimensions[r].height = 22

# Timeline data
timeline = [
    ["DAY 1", "Choose Blitz path.\nCommit in community.\nName Round 1 goal.", "Choose Prime Protocol path.\nCommit in community.\nName Round 1 goal.", "Final Fundability Lock:\n• Score verified\n• Utilization checked\n• Inquiries confirmed\n• Entity ready\n• Secondary bureaus clear\n→ GO DECISION", "4.1, 4.2"],
    ["DAY 1-2", "Cash loan seeded: $20-50K\n(0-2 biz days)\n→ FIRST FUNDING IN HAND", "—", "—", "4.3"],
    ["WEEK 1", "XXX institution apps: $10-20K x2-4\n(0-10 days w/ BRM)\n\nSecret strategy: $35-70K (instant)\n→ Personal + Business", "—", "Account Opening Sprint:\n• Chase account opened\n• Navy seeded (pledge loan)\n• BOA account opened\n• AMEX relationship started\n→ Seasoning clock starts", "4.3, 4.4"],
    ["WEEK 2", "Chase Business setup\n(special strategy + 2wk delay begins)", "—", "Bank Research begins (concurrent):\n• Online + forum research\n• Calling banks to confirm pulls\n• Product data extraction\n\nAccount Activity:\n• Transaction volume building\n• Daily balances growing\n• Money-in diversity", "4.5, 4.6"],
    ["WEEK 3", "AMEX apps submitted\n(1-9 week process begins)\n\nContinuing account activity", "—", "Research continues.\nAccounts seasoning.\nActivity patterns established.", "4.5, 4.6"],
    ["WEEK 4", "Chase Business FUNDED\n($35K-$150K)\n→ MAJOR WIN", "Research COMPLETE.\nTeam/support review.", "Strategy Review:\n• Full app sequence mapped\n• Bureau order locked\n• Contingency plans set\n→ APPLICATION PLAN FINALIZED", "4.7"],
    ["WEEK 5", "BofA apps submitted\n(1-2+ month delay begins)\n\nNF PL strategy initiated\n(1-2+ month delay begins)\n\nAdditional institution apps", "Round 1 Business-only apps\nsubmitted in strategic sequence.\n\nPersonal credit PRESERVED\nfor future rounds.\n→ APPLICATIONS LIVE", "Results documented.\nReal-time tracking.", "4.8 / 4.9"],
    ["WEEK 6-8", "AMEX results arriving\n($20K-$100K+)\n\nMore institution results\nflowing in", "Early results arriving.\nDocumenting everything.\nPreparing staggered apps.", "Ongoing monitoring.\nHandling approvals/denials.\nFollow-up actions.", "4.10"],
    ["MONTH 2-3", "BofA FUNDED ($25-100K)\nNF PL FUNDED ($50-100K)\n\nRemaining apps completing.\n→ BLITZ COMPLETE", "Chase Biz FUNDED ($35-150K)\nAMEX FUNDED ($20-100K+)\nBofA FUNDED ($25-100K)\n+ more institutions\n\nInquiry removal begins.\n→ ROUND 1 COMPLETE", "Staggered bank execution\nas seasoning windows open.", "4.10, 4.11"],
    ["MONTH 3-4", "Capital deployment.\nUtilization management.\n0% promo leverage.\n\n⚠️ NO new rounds for\n6-24+ months", "THE LOOP: Round 2 launched.\nSequence adjusted.\nCompounding begins.\n→ $100K becomes $300K+", "Capital deployment.\n1-3% monthly budget.\n6-12mo reserve account.", "4.11, 4.12"],
    ["MONTH 6+", "Still in cooldown.\nCannot launch new rounds.", "Round 3, 4, 5...\nRepeat every 1wk-1mo.\n→ $300K becomes $500K+", "—", "4.11"],
]

for i, row_data in enumerate(timeline):
    r = 5 + i
    ws1.row_dimensions[r].height = 95
    for j, val in enumerate(row_data):
        c = ws1.cell(row=r, column=j+1, value=val)
        c.font = normal_font
        c.alignment = wrap
        c.border = thin_border
    # Color coding
    ws1.cell(row=r, column=1).font = bold_font
    ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws1.cell(row=r, column=1).alignment = center_wrap
    if row_data[1] != "—":
        ws1.cell(row=r, column=2).fill = PatternFill("solid", fgColor=BLITZ_LIGHT)
    if row_data[2] != "—":
        ws1.cell(row=r, column=3).fill = PatternFill("solid", fgColor=PRIME_LIGHT)
    if row_data[3] != "—":
        ws1.cell(row=r, column=4).fill = PatternFill("solid", fgColor=SHARED_LIGHT)
    ws1.cell(row=r, column=5).font = small_font
    ws1.cell(row=r, column=5).alignment = center_wrap

# Summary row
r = 5 + len(timeline) + 1
ws1.merge_cells(f'A{r}:E{r}')
ws1.cell(row=r, column=1).value = "FUNDING POTENTIAL SUMMARY"
ws1.cell(row=r, column=1).font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws1.cell(row=r, column=1).fill = hdr_fill_purple
ws1.cell(row=r, column=1).alignment = Alignment(horizontal="center")
ws1.row_dimensions[r].height = 25

r += 1
summary = [
    ["Month 1 Potential", "$100K - $300K+", "$100K - $200K+", "", ""],
    ["Extended Potential", "$100K - $350K+ (months after)", "$50K - $250K+ (months after)", "", ""],
    ["Best Case", "$500K+ (if 2-3+ months patience)", "$500K+ (compound rounds)", "", ""],
    ["Repeat Eligibility", "❌ 6-24+ months cooldown", "✅ Repeat in 1 week - 1 month", "", ""],
    ["Cash Required", "None", "$5-10K+ reserved", "", ""],
    ["Credit Types Used", "Personal + Business", "Business only", "", ""],
    ["Future Rounds", "❌ Delayed 6-24+ months", "✅ Unlimited (The Loop)", "", ""],
    ["Recommended For", "One-time max extraction.\nDon't need funding again soon.", "Long-term funding strategy.\nMultiple rounds. Max total $.", "", ""],
]

for i, row_data in enumerate(summary):
    cr = r + i
    ws1.row_dimensions[cr].height = 30 if i < 7 else 45
    for j, val in enumerate(row_data):
        c = ws1.cell(row=cr, column=j+1, value=val)
        c.font = bold_font if j == 0 else normal_font
        c.alignment = wrap
        c.border = thin_border
    ws1.cell(row=cr, column=1).fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws1.cell(row=cr, column=2).fill = PatternFill("solid", fgColor=BLITZ_LIGHT)
    ws1.cell(row=cr, column=3).fill = PatternFill("solid", fgColor=PRIME_LIGHT)

# ======================================================================
# SHEET 2: BANK SEQUENCE
# ======================================================================
ws2 = wb.create_sheet("Bank Sequence")
ws2.sheet_properties.tabColor = BLITZ_ORANGE

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 16

ws2.merge_cells('A1:G1')
ws2['A1'].font = title_font
ws2['A1'].value = "BANK-SPECIFIC EXECUTION SEQUENCE"
ws2['A1'].alignment = Alignment(horizontal="center")

ws2.merge_cells('A3:G3')
ws2['A3'].font = Font(name="Arial", bold=True, color=BLITZ_ORANGE, size=13)
ws2['A3'].value = "THE BLITZ — V1a (Personal + Business, No Cash Required)"
ws2['A3'].fill = PatternFill("solid", fgColor=BLITZ_LIGHT)
ws2['A3'].alignment = Alignment(horizontal="center")

r = 4
blitz_headers = ["Step", "Institution / Strategy", "Funding Range", "Timeline", "In Blitz?", "In Prime?", "Type"]
for i, h in enumerate(blitz_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_header_row(ws2, r, 7, hdr_fill_blitz)

blitz_banks = [
    ["1", "Cash loan (seed relationships)", "$20K - $50K", "0-2 biz days", "✅", "❌", "Personal"],
    ["2", "XXX institution (x2-4 apps)", "$10K - $20K each", "0-10 days w/ BRM", "✅", "✅", "Business"],
    ["3", "Secret strategy", "$35K - $70K", "Instant", "✅", "✅ ($35K)", "Personal + Biz"],
    ["4", "Chase Business", "$35K - $150K", "2 week delay", "✅", "✅", "Business"],
    ["5", "AMEX", "$20K - $100K+", "1-9 week process", "✅", "✅ (Biz only)", "Personal + Biz"],
    ["6", "BofA Business", "$25K - $100K", "1-2+ month delay", "✅", "✅", "Business"],
    ["7", "Navy Federal PL strategy", "$50K - $100K", "1-2+ month delay", "✅", "❌", "Personal"],
    ["8", "Many more institutions", "$10K - $50K+ each", "0-3+ month delay", "✅", "✅ (Biz only)", "Personal + Biz"],
]

for i, row_data in enumerate(blitz_banks):
    cr = 5 + i
    ws2.row_dimensions[cr].height = 22
    for j, val in enumerate(row_data):
        c = ws2.cell(row=cr, column=j+1, value=val)
        c.font = normal_font
        c.alignment = wrap
        c.border = thin_border
    ws2.cell(row=cr, column=1).alignment = center_wrap
    ws2.cell(row=cr, column=5).alignment = center_wrap
    ws2.cell(row=cr, column=6).alignment = center_wrap
    if i % 2: 
        for j in range(1, 8):
            ws2.cell(row=cr, column=j).fill = PatternFill("solid", fgColor=BLITZ_LIGHT)

# Blitz totals
cr = 5 + len(blitz_banks)
ws2.cell(row=cr, column=1).value = ""
ws2.merge_cells(f'B{cr}:C{cr}')
ws2.cell(row=cr, column=2).value = "BLITZ TOTAL: $100K - $300K+ Month 1 | $100K - $350K+ Extended"
ws2.cell(row=cr, column=2).font = Font(name="Arial", bold=True, color=BLITZ_ORANGE, size=11)

# Prime Protocol section
cr += 2
ws2.merge_cells(f'A{cr}:G{cr}')
ws2.cell(row=cr, column=1).font = Font(name="Arial", bold=True, color=PRIME_BLUE, size=13)
ws2.cell(row=cr, column=1).value = "THE PRIME PROTOCOL — V2 (Business Only, $5-10K+ Cash Required)"
ws2.cell(row=cr, column=1).fill = PatternFill("solid", fgColor=PRIME_LIGHT)
ws2.cell(row=cr, column=1).alignment = Alignment(horizontal="center")

cr += 1
prime_headers = ["Step", "Institution / Strategy", "Funding Range", "Timeline", "In Prime?", "Repeat?", "Type"]
for i, h in enumerate(prime_headers, 1):
    ws2.cell(row=cr, column=i, value=h)
style_header_row(ws2, cr, 7, hdr_fill_prime)

prime_banks = [
    ["1", "XXX institution (x2-4 apps)", "$10K - $20K each", "0-5 biz days w/ BRM", "✅", "✅", "Business"],
    ["2", "Secret strategy", "$35K", "Instant", "✅", "✅", "Business"],
    ["3", "Chase Business", "$35K - $150K", "2 week delay", "✅", "✅", "Business"],
    ["4", "AMEX", "$20K - $100K+", "1-9 week process", "✅", "✅", "Business"],
    ["5", "BofA Business", "$25K - $100K", "1-2+ month delay", "✅", "✅", "Business"],
    ["6", "Many more institutions", "$10K - $50K+ each", "0-3+ month delay", "✅", "✅", "Business"],
]

for i, row_data in enumerate(prime_banks):
    rr = cr + 1 + i
    ws2.row_dimensions[rr].height = 22
    for j, val in enumerate(row_data):
        c = ws2.cell(row=rr, column=j+1, value=val)
        c.font = normal_font
        c.alignment = wrap
        c.border = thin_border
    ws2.cell(row=rr, column=1).alignment = center_wrap
    ws2.cell(row=rr, column=5).alignment = center_wrap
    ws2.cell(row=rr, column=6).alignment = center_wrap
    if i % 2:
        for j in range(1, 8):
            ws2.cell(row=rr, column=j).fill = PatternFill("solid", fgColor=PRIME_LIGHT)

rr = cr + 1 + len(prime_banks)
ws2.merge_cells(f'B{rr}:C{rr}')
ws2.cell(row=rr, column=2).value = "PRIME PROTOCOL TOTAL: $100K - $200K+ Month 1 | $50K - $250K+ Extended | Repeat in 1wk-1mo"
ws2.cell(row=rr, column=2).font = Font(name="Arial", bold=True, color=PRIME_BLUE, size=11)

# ======================================================================
# SHEET 3: PATH DECISION GUIDE
# ======================================================================
ws3 = wb.create_sheet("Path Decision Guide")
ws3.sheet_properties.tabColor = GREEN

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 35

ws3.merge_cells('A1:C1')
ws3['A1'].font = title_font
ws3['A1'].value = "WHICH PATH IS RIGHT FOR YOU?"
ws3['A1'].alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 35

ws3.merge_cells('A2:C2')
ws3['A2'].font = small_font
ws3['A2'].value = "Student-facing decision guide — used in Module 4.1 (Command Briefing)"
ws3['A2'].alignment = Alignment(horizontal="center")

r = 4
for i, h in enumerate(["QUESTION", "THE BLITZ", "THE PRIME PROTOCOL"], 1):
    ws3.cell(row=r, column=i, value=h)
fills3 = [hdr_fill_purple, hdr_fill_blitz, hdr_fill_prime]
for i, f in enumerate(fills3, 1):
    ws3.cell(row=r, column=i).fill = f
    ws3.cell(row=r, column=i).font = hdr_font
    ws3.cell(row=r, column=i).alignment = center_wrap
    ws3.cell(row=r, column=i).border = thin_border

decisions = [
    ["How fast do you want funding?", "As fast as possible.\nSame day to same week starts.", "Fast, but strategically paced.\n90-120 day full protocol."],
    ["Will you need funding again\nin the next 6-12 months?", "NO — I want one big extraction\nand I'm good for a while.", "YES — I want to compound\nand repeat funding rounds."],
    ["Are you ok using personal credit?", "YES — Personal + Business\nfor maximum extraction.", "NO — Business only.\nPersonal credit preserved."],
    ["Do you have $5-10K+ cash\nto keep reserved?", "Not required.", "YES — needed to seed\nbank relationships."],
    ["What's your M1 funding goal?", "$100K - $300K+\n(higher ceiling, one shot)", "$100K - $200K+\n(lower M1, but compounds)"],
    ["What's your total funding goal?", "$200K - $650K+\n(all in one launch)", "$300K - $500K+ per launch\n(unlimited launches)"],
    ["How much effort upfront?", "High intensity sprint.\nEverything at once.", "Strategic build.\nSteady compounding."],
    ["Risk tolerance?", "Higher — burning personal credit\nfor speed. One chance.", "Lower — preserving options.\nMultiple shots on goal."],
]

for i, row_data in enumerate(decisions):
    cr = 5 + i
    ws3.row_dimensions[cr].height = 45
    for j, val in enumerate(row_data):
        c = ws3.cell(row=cr, column=j+1, value=val)
        c.font = bold_font if j == 0 else normal_font
        c.alignment = wrap
        c.border = thin_border
    ws3.cell(row=cr, column=1).fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws3.cell(row=cr, column=2).fill = PatternFill("solid", fgColor=BLITZ_LIGHT)
    ws3.cell(row=cr, column=3).fill = PatternFill("solid", fgColor=PRIME_LIGHT)

# Recommendation row
cr = 5 + len(decisions)
ws3.row_dimensions[cr].height = 55
ws3.cell(row=cr, column=1).value = "RECOMMENDATION"
ws3.cell(row=cr, column=1).font = Font(name="Arial", bold=True, color=WHITE, size=11)
ws3.cell(row=cr, column=1).fill = hdr_fill_purple
ws3.cell(row=cr, column=1).alignment = center_wrap
ws3.cell(row=cr, column=1).border = thin_border

ws3.cell(row=cr, column=2).value = "Choose Blitz if you want\nmax funding NOW and\ndon't need rounds for 6-24mo."
ws3.cell(row=cr, column=2).font = blitz_font
ws3.cell(row=cr, column=2).fill = PatternFill("solid", fgColor=BLITZ_LIGHT)
ws3.cell(row=cr, column=2).alignment = center_wrap
ws3.cell(row=cr, column=2).border = thin_border

ws3.cell(row=cr, column=3).value = "⭐ RECOMMENDED ⭐\nChoose Prime Protocol for\nmax TOTAL funding over time.\nSmarter. Repeatable. Compounds."
ws3.cell(row=cr, column=3).font = prime_font
ws3.cell(row=cr, column=3).fill = PatternFill("solid", fgColor=PRIME_LIGHT)
ws3.cell(row=cr, column=3).alignment = center_wrap
ws3.cell(row=cr, column=3).border = thin_border

# ======================================================================
# SHEET 4: DFY ONLY DELAY COMPARISON
# ======================================================================
ws4 = wb.create_sheet("DFY ONLY vs DFY+DWY")
ws4.sheet_properties.tabColor = RED

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 22
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 22

ws4.merge_cells('A1:E1')
ws4['A1'].font = title_font
ws4['A1'].value = "WHY DFY+DWY BEATS DFY ONLY"
ws4['A1'].alignment = Alignment(horizontal="center")
ws4.row_dimensions[1].height = 35

ws4.merge_cells('A2:E2')
ws4['A2'].font = small_font
ws4['A2'].value = "Teaching moment in Module 4.1 — Why learning the system yourself = faster + cheaper"
ws4['A2'].alignment = Alignment(horizontal="center")

r = 4
dfy_headers = ["Step", "Action", "DFY + DWY", "DFY ONLY", "Delay Difference"]
for i, h in enumerate(dfy_headers, 1):
    ws4.cell(row=r, column=i, value=h)
style_header_row(ws4, r, 5, PatternFill("solid", fgColor=RED))

dfy_data = [
    ["1", "Cash loan seed", "0-2 biz days", "0-2 biz days\n+ wait for funds\n+ pay 10% fee", "+3-7 days"],
    ["2", "XXX institution apps", "0-10 days (concurrent)", "Must wait for Step 1\npayment cleared\n+ 0-10 days + pay 10%", "+1-2 weeks"],
    ["3", "Secret strategy", "Instant (concurrent)", "Must wait for Step 2\npayment cleared", "+1-2 weeks"],
    ["4", "Chase Business", "2 week delay (concurrent)", "Must wait for Step 3\npayment + 2 weeks", "+2-3 weeks"],
    ["5-8", "AMEX, BofA, Navy, etc.", "1-9 weeks (concurrent\nstacking)", "Sequential: 0-3+ MONTHS\nPER STEP between each", "+9-12+ months"],
]

for i, row_data in enumerate(dfy_data):
    cr = 5 + i
    ws4.row_dimensions[cr].height = 55
    for j, val in enumerate(row_data):
        c = ws4.cell(row=cr, column=j+1, value=val)
        c.font = normal_font
        c.alignment = wrap
        c.border = thin_border
    ws4.cell(row=cr, column=1).alignment = center_wrap
    ws4.cell(row=cr, column=1).font = bold_font
    ws4.cell(row=cr, column=3).fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    ws4.cell(row=cr, column=4).fill = PatternFill("solid", fgColor="FDEDEC")
    ws4.cell(row=cr, column=5).fill = PatternFill("solid", fgColor="FDEDEC")
    ws4.cell(row=cr, column=5).font = Font(name="Arial", bold=True, color=RED, size=10)

# Summary
cr = 5 + len(dfy_data) + 1
ws4.merge_cells(f'A{cr}:E{cr}')
ws4.cell(row=cr, column=1).value = "BOTTOM LINE"
ws4.cell(row=cr, column=1).font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws4.cell(row=cr, column=1).fill = PatternFill("solid", fgColor=RED)
ws4.cell(row=cr, column=1).alignment = Alignment(horizontal="center")

cr += 1
ws4.merge_cells(f'A{cr}:B{cr}')
ws4.cell(row=cr, column=1).value = "DFY + DWY (Learn + Execute)"
ws4.cell(row=cr, column=1).font = Font(name="Arial", bold=True, color=GREEN, size=11)
ws4.merge_cells(f'C{cr}:E{cr}')
ws4.cell(row=cr, column=3).value = "90-120 days to full funding launch. All steps concurrent. Max speed + max potential."
ws4.cell(row=cr, column=3).font = normal_font
ws4.cell(row=cr, column=3).fill = PatternFill("solid", fgColor=GREEN_LIGHT)
ws4.cell(row=cr, column=3).alignment = wrap
ws4.row_dimensions[cr].height = 30

cr += 1
ws4.merge_cells(f'A{cr}:B{cr}')
ws4.cell(row=cr, column=1).value = "DFY ONLY (No education)"
ws4.cell(row=cr, column=1).font = Font(name="Arial", bold=True, color=RED, size=11)
ws4.merge_cells(f'C{cr}:E{cr}')
ws4.cell(row=cr, column=3).value = "4+ weeks extra Month 1. 9-12+ months extra on extended phase. Sequential only. Less funding. More expensive."
ws4.cell(row=cr, column=3).font = normal_font
ws4.cell(row=cr, column=3).fill = PatternFill("solid", fgColor="FDEDEC")
ws4.cell(row=cr, column=3).alignment = wrap
ws4.row_dimensions[cr].height = 30

# Print settings
for ws in [ws1, ws2, ws3, ws4]:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = None

outpath = "./APW_Blitz_vs_Prime_Protocol_Roadmap.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
